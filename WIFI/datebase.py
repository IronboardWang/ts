import MySQLdb
import openpyxl
import pprint


conn = MySQLdb.connect(
    host='localhost',
    port = 3306,
    user="root",
    password="123456",
    db = "test", # 数据库名
    charset='utf8' # 字符
)
cur = conn.cursor()

cur.execute('SELECT * FROM test')
# # 读取一条
# row = cur.fetchone()
# print(row)
#读所有
# row = cur.fetchall()
# print(row)

# for i in range(cur.rowcount):
#      row = cur.fetchone()
#      print(row[1])

#分组读取
# for i in range(cur.rowcount):
#     rows = cur.fetchmany(100)
#     print(rows)
# 数据插入,从excel文件中
command = "INSERT INTO test VALUES (%s,%s)"
wb = openpyxl.load_workbook("transaction.xlsx")
sheet = wb["Sheet1"]
for row in range(1,sheet.max_row+1):
    #for column in range(1,sheet.max_column):
    id = (sheet.cell(row,1).value)
    name =sheet.cell(row,2).value
    cur.execute(command,(id,name))
# cur.execute("SELECT * FROM test")
cur = conn.cursor()

cur.execute('SELECT * FROM test')
row = cur.fetchall()
print(row)
cur.close()
conn.commit()
conn.close()